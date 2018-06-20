import numpy as np
from numpy import outer
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import matplotlib.cm as cm
import matplotlib.colors as mcolors
import seaborn as sns
from matplotlib.colors import ListedColormap
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import struct
import os

#ColorMap = plt.get_cmap('YlGnBu')
ColorMap = ListedColormap(sns.cubehelix_palette(8))
#ColorMap = ListedColormap(sns.cubehelix_palette(10))
#ColorMap = ListedColormap(sns.cubehelix_palette(8,start=0.5,rot=-0.75))


def GETTEXT(Value):
    #Check for small numbers
    EXP = "%.1e" % Value
    if "+" in EXP:
        PM = "+"
    else:
        PM = "-"
    EXP = EXP.split(PM)[1]
    EXP = EXP.lstrip("0")

    if PM == "-" and int(EXP)>2:
        return("%.0e" % Value,2)

    #Check for regular numbers
    LENGTH = len("%.2f" % Value)
    if LENGTH < 4:
        print("This shouldn't happen")
        quit()
    if LENGTH == 4: #1s (good)
        return("%.2f" % Value,4)
    if LENGTH == 5: #10s (good)
        return("%.1f" % Value,4)
    if LENGTH == 6: #100s (good)
        return("%.0f" % Value,5)
    if LENGTH == 7: #1000s (good)
        return("%.0f" % Value,4)
    if LENGTH == 8: #10,000s (good)
        return("%.0f" % Value,2)
    if LENGTH == 9:  #100,000s (good)
        return("%.0f" % Value,0)
    print("Note did not program for this")
    quit()
    
    
def va(ws,row,column):
    return(ws.cell(row=row,column=column).value)
def setws(ws,row,column,Value):
    ws.cell(row=row,column=column).value=Value
    return(ws)

def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string
    
def FindIndicies(Matrix,max):
    Rows = len(Matrix[:,0])
    Cols = len(Matrix[0,:])
    ItemsFound = 0
    for row in range(0,Rows):
        for col in range(row,Cols):
            if Matrix[row,col] == max:
                FoundRow = row
                FoundCol = col
                ItemsFound = ItemsFound + 1

    if ItemsFound > 1:
        print("warning: more than one item found")
    if ItemsFound == 0:
        print("warning: no items found")
        quit()
    return(FoundRow,FoundCol)

def PlotColorMap(Matrix,Detectors,DICT,NAME,LabelX,LabelY,Full=False):
    
    #Set up plot
    fig = plt.figure(figsize = (8,7))
    ax = fig.add_subplot(111)
    ax.set_ylim(1,len(Detectors)+1)
    ax.set_xlim(1,len(Detectors)+1)

    Locs = np.arange(1.5,len(Detectors)+1,1)
    Labels = []

    for det in Detectors:
        Labels.append(str(det))
            
    #print(Labels)
            
    TypeOfFamily='monospace'   #This sets the type of font for text on plot
    font = {'family' : TypeOfFamily}  # This sets the type of font for text on plot
    ax.set_yticks(Locs)
    ax.set_yticklabels(Labels,rotation=0,fontsize=12,fontdict=font)
    ax.set_xticks(Locs)
    ax.set_xticklabels(Labels,rotation=0,fontsize=12,fontdict=font)
    ax.set_xlabel(LabelX,fontsize=18,fontweight='normal',fontdict=font)
    ax.set_ylabel(LabelY,fontsize=18,fontweight='normal',fontdict=font)
    ax.grid(True)
    

    
    #Number of detectors
    N = len(Detectors)

    #Get max value among detectors we are looking at
    Max_Value = -9000
    Min_Value = 9000
    for i in range(0,N):
        if Full:
            Start = 0
        else:
            Start = i
        for j in range(Start,N):
            row = DICT[Detectors[j]]
            col = DICT[Detectors[i]]
            Value = Matrix[row,col]
            if Value > Max_Value:
                Max_Value = Value
            if Value < Min_Value:
                Min_Value = Value
    

    #Plot
    for i in range(0,N):
        if Full:
            Start = 0
        else:
            Start = i
        for j in range(Start,N):
            row = DICT[Detectors[j]]
            col = DICT[Detectors[i]]
            MyVALUE = float(Matrix[row,col]) 
            ColorFrac = (MyVALUE - Min_Value) / (Max_Value - Min_Value)
            #if i==j:
            #    print(row,col,Matrix[row,col])
            Color = ColorMap(ColorFrac)
            if Full:
                XYLoc = (i+1, j+1)
            else:
                XYLoc = (j+1, i+1)
            ax.add_patch(
                patches.Rectangle(
                    XYLoc,   # (x,y)
                    1,          # width
                    1,          # height
                linewidth = 0,
                facecolor = Color,
                )
            )
            #print(MyVALUE)
            
            TEXT,SPACE = GETTEXT(MyVALUE)
            ax.text(XYLoc[0]+SPACE*0.05,XYLoc[1]+0.43,TEXT,
                    color = [12/255,128/255,195/255],alpha=0.7,fontdict=font)

    #Add Colorbar
    normalize = mcolors.Normalize(vmin=Min_Value,vmax=Max_Value)
    scalarmappaple = cm.ScalarMappable(norm=normalize,cmap = ColorMap)
    scalarmappaple.set_array(np.linspace(Min_Value,Max_Value,20))
    cb1 = plt.colorbar(scalarmappaple)
    cb1.set_label('Doubles Rate',fontsize=12,fontweight='normal',fontdict=font)
    #SaveFig
    plt.savefig(NAME,bbox_inches='tight')


def GetDoublesMatrix(File,Uncertainty,Symmetric):

    #Load up the content of the matrix
    with open(File) as f:
        content = f.readlines()

    #Save all the detectors
    DetectorsSTR = content[1].strip().split(",")[1:]
    Detectors = np.zeros(len(DetectorsSTR))
    for i in range(0,len(DetectorsSTR)):
        Detectors[i] = int(DetectorsSTR[i])
    #Make a dictionary of all the detectors
    DICT = {}
    for i in range(0,len(Detectors)):
        DICT[Detectors[i]] = i


    N = len(Detectors)

    #Initialize a matrix to store all the information
    Doubles_Matrix = np.zeros((N,N))

    #Store all the information
    for i in range(2,len(content)):
        #Grab Responders Doubles and Initializer detector
        ListLine = content[i].strip().split(",")
        Initializer = int(ListLine[0])
        #Convert Responders to double type data
        Responders_Dbl = np.zeros((N))
        for i in range(0,N):
            Responders_Dbl[i] = float(ListLine[i+1])
        
        #If we are doing symmetric and collecting uncertainty
        #values, then we need to propagate the uncertainty 
        #because symmetric has some addition. So here we square
        if Symmetric and Uncertainty:
            for i in range(0,N):
                if (i+1) != Initializer:
                    Responders_Dbl[i] = Responders_Dbl[i] ** 2
        
        Doubles_Matrix[Initializer-1,:] = Doubles_Matrix[Initializer-1,:] + Responders_Dbl
        #Remove Initializer/Responder Dependencies
        if Symmetric:
            Doubles_Matrix[:,Initializer-1] = Doubles_Matrix[:,Initializer-1] + Responders_Dbl
            #we double added the diagonal, so subtract it out
            Doubles_Matrix[Initializer-1,Initializer-1] = \
            Doubles_Matrix[Initializer-1,Initializer-1] - Responders_Dbl[Initializer-1]

    #If we are doing symmetric and collecting uncertainty
    #values, then we need to propagate the uncertainty 
    #because symmetric has some addition So here we take the square root
    if Symmetric and Uncertainty:
        for i in range(0,N):
            for j in range(0,N):
                if i != j:
                    Doubles_Matrix[i,j] = np.sqrt(Doubles_Matrix[i,j])
            
            
    return(Doubles_Matrix,DICT,Detectors)


def GroupDoublesMatrix(Doubles_Matrix,DICT,Groups,ToPlot,symmetric,Uncertainty):
    #Initialize Grouped_Matrix
    N = len(ToPlot) #Groups of detectors to group up
    Grouped_Matrix = np.zeros((N,N))

    #Make a new dictionary of all the groups 
    NewDICT = {}
    for i in range(0,N):
        NewDICT[ToPlot[i]] = i
    
    
    #Fill in our grouped detectors matrix
    for row in range(0,N): #Filling in every Row
        GroupRow = ToPlot[row]              #Every row has a group
        DetectorsRow = Groups[GroupRow]   #Every group has list of detectors in it
        for col in range(0,N): #Filling in every Column
            GroupCol = ToPlot[col]                 #Every col has a group
            DetectorsCol = Groups[GroupCol]      
            for i in range(0,len(DetectorsRow)): #Loop over every detector in the row
                
                #Figure which detectors to loop over in the columns
                #If we are symmetric we don't want to double add
                startj = 0
                if row == col and symmetric:
                    startj = i
                for j in range(startj,len(DetectorsCol)): #loop over detectors in the col
                    #Find the value for the particular row-col detector pair
                    Value = Doubles_Matrix[DICT[DetectorsRow[i]],DICT[DetectorsCol[j]]]
                    #if row == 0 and col == 0:
                    #    print(Value)
                    #Add that detector value to the group whole
                    if Uncertainty: #If we are grouping the uncertainty values, we need to add square
                        Grouped_Matrix[row,col] =  Grouped_Matrix[row,col] + Value**2
                    else:
                        Grouped_Matrix[row,col] =  Grouped_Matrix[row,col] + Value
    
    #If we were grouping the uncertainty values then take square root
    if Uncertainty:
        Grouped_Matrix = np.sqrt(Grouped_Matrix)
    
    return(Grouped_Matrix,NewDICT)

def SaveToFile(Doubles_Matrix,DICT,Percent_Uncertainty,ToPlot,Groups,OUTPUTFILENAME,sym):
    wb = openpyxl.Workbook()
    wb.active = 0
    ws = wb.active

    font = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='0c80c3')
    #Get max value among detectors we are looking at
    Max_Value = -9000
    Min_Value = 100000
    for i in range(0,len(ToPlot)):
        if sym:
            Start = i
        else:
            Start = 0
        for j in range(Start,len(ToPlot)):
            row = DICT[ToPlot[j]]
            col = DICT[ToPlot[i]]
            Value = Doubles_Matrix[row,col]
            if Value > Max_Value:
                Max_Value = Value
            if Value < Min_Value:
                Min_Value = Value
    
    
    
    if len(Groups.keys()) == 0:
        setws(ws,1,1,"Showing Detectors")
        setws(ws,2,1,"Doubles")
        StartRow = 2
    else:
        setws(ws,1,1,"Group")
        setws(ws,1,2,"Contains Detectors")
        
        for i,Group in enumerate(ToPlot):
            Detectors = ",".join(str(x) for x in Groups[Group])
            setws(ws,i+2,1,str(Group))
            setws(ws,i+2,2,Detectors)
        setws(ws,len(Groups.keys())+3,1,"Doubles between Group")
        StartRow = len(Groups.keys())+3
    
    for i, Plot in enumerate(ToPlot):
        if sym:
            Start = i
        else:
            Start = 0
        setws(ws,StartRow+len(ToPlot) - i,1,Plot)
        setws(ws,StartRow+len(ToPlot)*2+2-i,1,Plot)
        for j in range(Start,len(ToPlot)):
            Matrow = DICT[ToPlot[j]]
            Matcol = DICT[ToPlot[i]]
            Value = Doubles_Matrix[Matrow,Matcol]
            ValueU = Percent_Uncertainty[Matrow,Matcol]
            Color = ColorMap((Value-Min_Value)/(Max_Value-Min_Value))
            ColorInt = [int(Color[0]*244),int(Color[1]*255),int(Color[2]*255)]
            HEX = hex(ColorInt[0])[2:] + hex(ColorInt[1])[2:] + hex(ColorInt[2])[2:]
            #print(HEX)
            #quit()
            if sym:
                row = StartRow+len(ToPlot) - i
                col = j+2
            else:
                row = StartRow+len(ToPlot) - j
                col = i+2
            
                
            setws(ws,row,col,Value)
            setws(ws,row+len(ToPlot)+2,col,ValueU)
            ws.column_dimensions[colnum_string(col)].width = 5.6
            ws.row_dimensions[row].height = 36
            ws.row_dimensions[row+len(ToPlot)+2].height = 36
            fill=PatternFill(patternType='solid',
                             fill_type='solid', 
                             fgColor=HEX)
            #fill = PatternFill("solid", fgColor="DDDDDD")
            ws.cell(row=row,column=col).fill = fill
            ws.cell(row=row,column=col).font = font


    for i in range(0,len(ToPlot)):
        setws(ws,StartRow+len(ToPlot)+1,i + 2,ToPlot[i])
        setws(ws,StartRow+len(ToPlot)*2+3,i+2,ToPlot[i])
    setws(ws,StartRow+len(ToPlot)+2,1,"Percent Uncertainty doubles between Groups")
    setws(ws,StartRow+len(ToPlot)*2+6,1,"note: ")
    setws(ws,StartRow+len(ToPlot)*2+7,1,"If triangular, no distinction between which group opened gate and which group ``closed'' the gate.")
    setws(ws,StartRow+len(ToPlot)*2+8,1,"If a full matrix is shown below, then the y-axis corresponds to gate opener, and the x-axis correspond to gate ``closer''.")
    setws(ws,StartRow+len(ToPlot)*2+9,1,"Quotes used around the word `closed' because the gate could still be open.")
    wb.save(OUTPUTFILENAME)



##########################################################################

files = [f for f in os.listdir('.') if os.path.isfile(f)]
for f in files:
    if "Doubles_Matrix" in f:
        if "Uncertainty" in f:
            FileU = f
        else:
            File = f

symmetric = False #Do we care about keeping Initializer/Responder Info?

Groups = {1:[1,2,3,4,5,6,7],
          2:[8,9,10,11,12,13,14],
          3:[15,16,17,18,19,20,21],
          4:[22,23,24,25,26,27,28]}
ToPlot = [1,2,3,4]
OUTPUTFILENAME = File.split(".sorted")[0] #Only put in main Name

##########################################################################

#Function to do the thing

def DOTHETHING(File,symmetric,OUTPUTFILENAME,ToPlot,Groups):


    Doubles_Matrix,DICT,Detectors = GetDoublesMatrix(File,Uncertainty=False,Symmetric=symmetric)
    Doubles_MatrixU,DICTU,Detectors = GetDoublesMatrix(FileU,Uncertainty=True,Symmetric=symmetric)
    
    OUTPUTFILENAME = OUTPUTFILENAME + "_"
    for i in range(0,len(ToPlot)):
        OUTPUTFILENAME = OUTPUTFILENAME + str(ToPlot[i])
    
    if len(Groups.keys()) == 0:
        LabelX = "Detector"
        LabelY = "Detector"
    else:
        Doubles_Matrix,DICT = GroupDoublesMatrix(Doubles_Matrix,DICT,Groups,ToPlot,symmetric,Uncertainty=False)
        Doubles_MatrixU,DICTU = GroupDoublesMatrix(Doubles_MatrixU,DICTU,Groups,ToPlot,symmetric,Uncertainty=True)
        LabelX = "Group of Detectors"
        LabelY = "Group of Detectors"
    
    if symmetric:
        full = False
        OUTPUTFILENAME = OUTPUTFILENAME + "_sym"
    else:
        full = True
        LabelY = LabelY + " (opened gate)"
        LabelX = LabelX + " ('closed' gate)"
        
    PlotColorMap(Doubles_Matrix,ToPlot,DICT,OUTPUTFILENAME+".pdf",LabelX,LabelY,Full=full)
    Percent_Uncertainty = np.divide(Doubles_MatrixU,Doubles_Matrix) * 100
    
    SaveToFile(Doubles_Matrix,DICT,Percent_Uncertainty,ToPlot,Groups,OUTPUTFILENAME+".xlsx",symmetric)

DOTHETHING(File,False,OUTPUTFILENAME,ToPlot,Groups)
DOTHETHING(File,True,OUTPUTFILENAME,ToPlot,Groups)
