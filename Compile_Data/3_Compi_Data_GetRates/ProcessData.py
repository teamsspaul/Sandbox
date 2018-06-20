############################################
############ Import Packages ###############
############################################

from os import listdir
from os.path import isfile,join
from openpyxl import load_workbook
from uncertainties import ufloat
import numpy as np
import matplotlib
import matplotlib.pyplot as plt

############################################
################ Change Here ###############
############################################

GW = "24"
PD = "4"
NewParameters = ["S.DW"+GW+".PD"+PD,"SErr.DW"+GW+".PD"+PD,
                 "D.DW"+GW+".PD"+PD,"DErr.DW"+GW+".PD"+PD,
                 "T.DW"+GW+".PD"+PD,"TErr.DW"+GW+".PD"+PD,
                 "Q.DW"+GW+".PD"+PD,"QErr.DW"+GW+".PD"+PD,
                 "P.DW"+GW+".PD"+PD,"PErr.DW"+GW+".PD"+PD,
                 "tau1.Kat.GW"+GW+".PD"+PD,"tau2.Kat.GW"+GW+".PD"+PD,
                 "GF2.Kat.GW"+GW+".PD"+PD,"GF3.Kat.GW"+GW+".PD"+PD,
                 "AS.Kat.GW"+GW+".PD"+PD,"RAS.Kat.GW"+GW+".PD"+PD,
                 "AD.Kat.GW"+GW+".PD"+PD,"RAD.Kat.GW"+GW+".PD"+PD,
                 "AT.Kat.GW"+GW+".PD"+PD,"RAT.Kat.GW"+GW+".PD"+PD,
                 "AQ.Kat.GW"+GW+".PD"+PD,"RAQ.Kat.GW"+GW+".PD"+PD,
                 "AP.Kat.GW"+GW+".PD"+PD,"RAP.Kat.GW"+GW+".PD"+PD,
                 "t2dg.GW"+GW+".PD"+PD,"SumRa.GW"+GW+".PD"+PD,"SumA.GW"+GW+".PD"+PD
                ]
NewParameterDefinitions = [
    "Singles rate with a specific GW and PD microsecond bin width",
    "Above parameter uncertainty in percent",
    "Doubles rate with a specific GW and PD microsecond bin width",
    "Above parameter uncertainty in percent",
    "Triples rate with a specific GW and PD microsecond bin width",
    "Above parameter uncertainty in percent",
    "Quads rate with a specific GW and PD microsecond bin width",
    "Above parameter uncertainty in percent",
    "Pents rate with a specific GW and PD microsecond bin width",
    "Above parameter uncertainty in percent",
    "Tau 1 calculated with FT with specific GW and PD",
    "Tau 2 calculated with FT with specific GW and PD",
    "GF doubles calculated with FT with specific GW and PD",
    "GF triples calculated with FT with specific GW and PD",
    "Sum parameter","Sum parameter","Sum parameter","Sum parameter","Sum parameter",
    "Sum parameter","Sum parameter","Sum parameter","Sum parameter","Sum parameter",
    "Two gate doubles?","Sum parameter","Sum parameter"
]

if len(NewParameters) != len(NewParameterDefinitions):
    print("Make sure you have the same amount of parameters and definitions")
    quit()

def OptionalWork(File):
    return()
    content = open(File).readlines()
    GateWidth = np.zeros((len(content)))
    Doubles = np.zeros((len(content)))
    Uncertainty = np.zeros((len(content)))
    for i in range(0,len(content)):
        line = content[i].split("\n")[0]
        GateWidth[i] = float(line.split(",")[0])
        Doubles[i] = float(line.split(",")[1])
        Uncertainty[i] = float(line.split(",")[2])
    GateWidth,Uncertainty = (list(t) for t in zip(*sorted(zip(GateWidth,Uncertainty))))
    fig, ax = plt.subplots()
    ax.plot(GateWidth, Uncertainty,'b')
    ax.set(xlabel='Gate Width', ylabel='Percent Uncertainty')
    ax.grid()
    fig.savefig(File.replace(".txt",".pdf"))
def GetValues(File):
    content = open(File).readlines()
    print(content[-1])
    quit()
    GateWidth = np.zeros((len(content)))
    Doubles = np.zeros((len(content)))
    Uncertainty = np.zeros((len(content)))
    for i in range(0,len(content)):
        line = content[i].split("\n")[0]
        GateWidth[i] = float(line.split(",")[0])
        Doubles[i] = float(line.split(",")[1])
        Uncertainty[i] = float(line.split(",")[2])
    min = 100
    for i in range(0,len(GateWidth)):
        #print("%.0f" % (GateWidth[i]*10**6) + " %.3f" % Uncertainty[i])
        if Uncertainty[i] < min:
            min = Uncertainty[i]
            MinIndex = i
    
    MinGateWidth = GateWidth[MinIndex]*10**6
    MinDoubles = Doubles[MinIndex]
    MinUncertainty = Uncertainty[MinIndex]
    
    return([MinGateWidth,MinDoubles,MinUncertainty])

############################################
################# End change ###############
############################################

############################################
################ Functions #################
############################################
def va(ws,row,column):
    return(ws.cell(row=row,column=column).value)
def setws(ws,row,column,Value):
    ws.cell(row=row,column=column).value=Value
    return(ws)
def v(Value):
    return(Value.nominal_Value)
def s(Value):
    return(Value.std_dev)
def GetParameterList(ws):
    List = []
    for i in range(2,10000):
        Value = va(ws,1,i)
        if Value is not None:
            List.append(Value)
    return(List)
def GetAssayDict(ws):
    AssyDict = {}
    for i in range(2,10000):
        Value = str(va(ws,i,1))
        if Value is not None:
            AssyDict[Value] = i
    return(AssyDict)
def GetIndexes(NewParameters,PList):
    Indexes = []
    Added = 0
    for i in range(0,len(NewParameters)):
        Param = NewParameters[i]
        Have = False
        for j in range(0,len(PList)):
            pitem = PList[j]
            if pitem == Param:
                Have = True
                Indexes.append(j+2)
                break
        if not Have:
            Indexes.append(2+len(PList)+Added)
            Added = Added + 1
    return(Indexes)
def UpdateParameters(wsP,NewParameters,Indexes):
    for i in range(0,len(NewParameters)):
        wsP = setws(wsP,1,Indexes[i],NewParameters[i])
    return(wsP)
def UpdateKeys(wsK,NewParameters,Indexes,NewParameterDefinitions):
    for i in range(0,len(NewParameters)):
        Str = NewParameters[i] + " = " + NewParameterDefinitions[i]
        wsK = setws(wsK,7+Indexes[i],1,Str)
    return(wsP)
############################################
################# Program ##################
############################################

#Grab all the output files
dir = "Collected_Files" 
Files = [join(dir,f) for f in listdir(dir) if isfile(join(dir,f)) and "ReExcel" in f]

#Open up the excel sheet
ExcelFile = "../30_Compiled_Data/FT_Data.xlsx"
wb = load_workbook(ExcelFile)
wsP = wb["Parameters"]
wsK = wb["Key"] 
#Grab current parameters and assys being tracked in the excel sheet
PList = GetParameterList(wsP)
AssyDict = GetAssayDict(wsP)

#Determine associated indexes for where parameter values should be 
#stored, if the parameters already exists, then the program will 
#overwrite the values
Indexes = GetIndexes(NewParameters,PList)

#Update the excel sheet main page and key page
wsP = UpdateParameters(wsP,NewParameters,Indexes)
wsK = UpdateKeys(wsK,NewParameters,Indexes,NewParameterDefinitions)

#Fill in the Values
for i in range(0,len(Files)):
    File = Files[i]
    print(File)
    if "\\" in File:
        Assy = File.split("\\")[1].split(".")[0].split("_")
    elif "/" in File:
        Assy = File.split("/")[1].split(".")[0].split("_")
    else:
        print("WTF");quit()
    if len(Assy) ==2:
        Assy = Assy[1].split("-")[0]
    else:
        Assy = "_".join(Assy[1:]).split("-")[0]
    Row = AssyDict[Assy]
    OptionalWork(File)
    Values = GetValues(File)
    if len(Values) != len(NewParameters):
        print("Length of returned values and parameters inconsistent")
        quit()
    for j in range(0,len(NewParameters)):
        wsP = setws(wsP,Row,Indexes[j],Values[j])

wb.save(ExcelFile)


